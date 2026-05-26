import streamlit as st
import pandas as pd
import io
import openpyxl
import re
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.formatting.rule import FormulaRule
from datetime import date, datetime

st.set_page_config(page_title="WIP Summary Consolidator Pro v54", layout="wide")

def extract_area_from_filename(filename):
    match = re.search(r"Open Orders List\s+(.*?)\s+\d+", filename, re.IGNORECASE)
    area = "Unknown"
    if match: 
        area = match.group(1).strip()
    else:
        parts = filename.replace('.xlsx', '').split(' ')
        area = parts[3] if len(parts) >= 4 else "Unknown"
    
    if area.lower() == "hyderabad": return "HYD"
    if area.upper() == "A.P": return "Nellore"
    return area

def to_numeric(val):
    if val is None or pd.isna(val): return ""
    if isinstance(val, (int, float)): return val
    s = str(val).strip().replace(',', '')
    if not s or s.lower() in ['nan', 'none']: return ""
    
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return str(val).strip()

def clean_date_string(val):
    if val is None or pd.isna(val): return None
    if isinstance(val, (datetime, date)):
        if hasattr(val, 'date'): return val.date()
        return val
    s = str(val).strip().replace('\t', '').replace('\n', '')
    if not s or s.lower() in ['nan', 'none', '']: return None
    try:
        return pd.to_datetime(s).date()
    except:
        return s

def parse_excel_wip(file_obj, filename):
    try:
        df_raw = pd.read_excel(file_obj, header=None)
    except Exception as e:
        st.error(f"Error reading {filename}: {e}")
        return []

    data = []
    today = date.today()
    area = extract_area_from_filename(filename)

    header_idx = -1
    col_map = {}
    for i in range(min(50, len(df_raw))):
        row_list = [str(x).strip() for x in df_raw.iloc[i]]
        if 'Pending Days' in row_list:
            header_idx = i
            for j, val in enumerate(row_list):
                if val and str(val) != 'nan':
                    col_map[val] = j
            break
    
    if header_idx == -1: return []

    curr_dname = "Unknown"
    curr_dealer = "Unknown"
    i = header_idx + 1
    
    while i < len(df_raw):
        row = df_raw.iloc[i]
        if str(row[0]).strip() == 'Dealer':
            row_list_clean = [str(x).strip() for x in row.values if pd.notna(x) and str(x).strip() not in ['', 'nan']]
            if len(row_list_clean) >= 3:
                curr_dealer = row_list_clean[2]
                # Extract Dname directly from the text before "|" of Dealer Name if available
                curr_dname = curr_dealer.split('|')[0].strip() if '|' in curr_dealer else row_list_clean[1].split('.')[0].strip()
            elif len(row_list_clean) == 2:
                curr_dealer = row_list_clean[1]
                curr_dname = curr_dealer.split('|')[0].strip() if '|' in curr_dealer else row_list_clean[1].split('.')[0].strip()
            i += 1
            continue
            
        try:
            val_0 = row[0]
            if not pd.isna(val_0) and (isinstance(val_0, (int, float)) or str(val_0).isdigit()):
                def get_v(label):
                    idx = col_map.get(label)
                    return row[idx] if idx is not None else None

                ord_no = str(get_v('Ord.No.')).split('.')[0].strip()
                cr_dt_raw = clean_date_string(get_v('Cr.Dt'))
                
                calc_pending = ""
                if cr_dt_raw and isinstance(cr_dt_raw, date):
                    calc_pending = int((today - cr_dt_raw).days)
                else:
                    try: calc_pending = int(float(str(val_0)))
                    except: calc_pending = val_0

                record = {
                    "SNO.": "",
                    "Pending Days": calc_pending,
                    "D.Code &JC NO.": f"{curr_dname}{ord_no}",
                    "Dname": curr_dname,
                    "Area": area,
                    "Dealer Name": curr_dealer,
                    "Req. Delv. Dt": clean_date_string(get_v('Req. Delv. Dt')),
                    "Cr.Dt": cr_dt_raw,
                    "Total": get_v('Total'),
                    "PartsTotal": get_v('PartsTotal'),
                    "Ord.No.": ord_no,
                    "Ord.Ty.": get_v('Ord.Ty.'),
                    "Regn.No": get_v('Regn.No'),
                    "Chassis/SL No.": get_v('Chassis/SL No.'),
                    "Notes": "",
                    "Cust.No": get_v('Cust.No'),
                    "Cust Name": get_v('Cust Name'),
                    "Remarks": "",
                    "Category": "",
                    "Invoice no.": "",
                    "Date": "",
                    "Status": "Open"
                }
                
                if i + 1 < len(df_raw):
                    next_row_vals = [str(x).strip() for x in df_raw.iloc[i+1]]
                    if 'Notes' in next_row_vals:
                        row_notes = df_raw.iloc[i+1]
                        for k in range(min(21, len(df_raw.columns)), len(df_raw.columns)):
                            if not pd.isna(row_notes[k]) and str(row_notes[k]).strip() not in ["", "Notes"]:
                                record['Notes'] = row_notes[k]
                                break
                        i += 1
                data.append(record)
        except: pass
        i += 1
    return data

def process_paycode_report(file_obj):
    if not file_obj: return {}
    try:
        df_raw = pd.read_excel(file_obj, header=None)
        header_row = -1
        ord_idx = -1
        inv_idx = -1
        date_idx = -1
        
        for i in range(min(25, len(df_raw))):
            row_vals = [str(x).strip() for x in df_raw.iloc[i].values]
            if "Ord.ID" in row_vals:
                header_row = i
                ord_idx = row_vals.index("Ord.ID")
                
                for idx, val in enumerate(row_vals):
                    v_lower = val.lower()
                    if "invoicedt" in v_lower or "invoice dt" in v_lower or "inv.dt" in v_lower:
                        date_idx = idx
                    elif "inv.no" in v_lower or "invoice" in v_lower or "inv no" in v_lower:
                        inv_idx = idx
                break
                
        if header_row != -1 and inv_idx > 0:
            left_col_count = df_raw.iloc[header_row+1:, inv_idx-1].dropna().count()
            curr_col_count = df_raw.iloc[header_row+1:, inv_idx].dropna().count()
            if left_col_count > curr_col_count:
                inv_idx = inv_idx - 1
                
        if header_row != -1 and date_idx > 0:
            left_dt_count = df_raw.iloc[header_row+1:, date_idx-1].dropna().count()
            curr_dt_count = df_raw.iloc[header_row+1:, date_idx].dropna().count()
            if left_dt_count > curr_dt_count:
                date_idx = date_idx - 1

        paycode_dict = {}
        if header_row != -1 and ord_idx != -1:
            for i in range(header_row + 1, len(df_raw)):
                ord_raw = df_raw.iloc[i, ord_idx]
                if pd.isna(ord_raw): continue
                
                ord_val = str(ord_raw).split('.')[0].strip()
                if not ord_val or ord_val.lower() in ['nan', 'ord.id', 'none']:
                    continue
                    
                inv_val = None
                if inv_idx != -1:
                    raw_inv = df_raw.iloc[i, inv_idx]
                    if not pd.isna(raw_inv):
                        i_str = str(raw_inv).split('.')[0].strip()
                        if i_str.lower() not in ['nan', 'none', '']:
                            inv_val = i_str
                            
                dt_val = None
                if date_idx != -1:
                    raw_dt = df_raw.iloc[i, date_idx]
                    if not pd.isna(raw_dt):
                        dt_val = clean_date_string(raw_dt)
                            
                if ord_val not in paycode_dict:
                    paycode_dict[ord_val] = []
                    
                if inv_val or dt_val:
                    paycode_dict[ord_val].append({"invoice": inv_val, "date": dt_val})
                        
        return paycode_dict
    except Exception as e:
        pass
    return {}

def apply_rich_remarks(text):
    if not text or not isinstance(text, str) or text.strip() == "" or text.lower() == 'nan': return ""
    red_bold = InlineFont(color="FFFF0000", b=True)
    normal = InlineFont()
    rt = CellRichText()
    parts = text.split('-', 1)
    rt.append(TextBlock(red_bold, parts[0]))
    if len(parts) > 1:
        suffix = "-" + parts[1]
        if "Job completed" in suffix:
            sub_parts = suffix.split("Job completed")
            for idx, p in enumerate(sub_parts):
                rt.append(TextBlock(normal, p))
                if idx < len(sub_parts) - 1: rt.append(TextBlock(red_bold, "Job completed"))
        else: rt.append(TextBlock(normal, suffix))
    return rt

st.title("📊 WIP Summary Consolidator Pro v54")

with st.sidebar:
    st.header("Files")
    summary_file = st.file_uploader("1. Existing Summary File (Optional)", type=["xlsx"])
    open_order_files = st.file_uploader("2. New Open Order Lists", type=["xlsx"], accept_multiple_files=True)
    paycode_file = st.file_uploader("3. Paycodewise Report (Optional)", type=["xlsx"])

if st.button("Generate Final Report"):
    if not open_order_files and not summary_file:
        st.warning("Please upload files to proceed.")
    else:
        existing_data = []
        wb = None
        target_name = "Master Data"
        
        std_headers = ['SNO.', 'Pending Days', 'D.Code &JC NO.', 'Dname', 'Area', 'Dealer Name', 'Req. Delv. Dt', 'Cr.Dt', 'Total', 'PartsTotal', 'Ord.No.', 'Ord.Ty.', 'Regn.No', 'Chassis/SL No.', 'Notes', 'Cust.No', 'Cust Name', 'Closing Date', 'Remarks', 'Category', 'Invoice no.', 'Date', 'Status']

        if summary_file:
            wb = openpyxl.load_workbook(summary_file, keep_links=True)
            target_name = "Master Data" if "Master Data" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[target_name]

            data_rows = list(ws.iter_rows(values_only=True))
            if data_rows:
                header_ws = data_rows[0]
                col_map_ws = {name: i for i, name in enumerate(header_ws)}
                
                for row in data_rows[1:]:
                    if all(val is None or str(val).strip() == '' for val in row):
                        continue
                        
                    d = {header_ws[i]: val for i, val in enumerate(row)}
                    
                    ord_key = next((k for k in d.keys() if str(k).strip().upper() in ["ORD.NO.", "ORD NO"]), None)
                    if not ord_key or not d.get(ord_key) or str(d[ord_key]).strip().lower() in ['nan', 'none', '']:
                        continue

                    if d.get('Area') == "A.P": d['Area'] = "Nellore"
                    if d.get('Area') == "Hyderabad": d['Area'] = "HYD"
                    
                    for k in ['Req. Delv. Dt', 'Cr.Dt', 'Closing Date', 'Date']:
                        if d.get(k): d[k] = clean_date_string(d[k])

                    pd_key = next((k for k in d.keys() if str(k).strip().upper() in ["PENDING DAYS", "PENDING"]), None)
                    if pd_key and d[pd_key] is not None:
                        try: d[pd_key] = int(float(str(d[pd_key])))
                        except: pass
                    existing_data.append(d)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Master Data"
            col_map_ws = {name: i for i, name in enumerate(std_headers)}
            
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col_idx, col_name in enumerate(std_headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font

        new_data_raw = []
        if open_order_files:
            for f in open_order_files:
                new_data_raw.extend(parse_excel_wip(f, f.name))
        
        def get_ord(r):
            k = next((key for key in r.keys() if str(key).strip().upper() in ["ORD.NO.", "ORD NO"]), None)
            return str(r.get(k, '')).strip() if k else ''

        existing_ids = {get_ord(r) for r in existing_data if get_ord(r)}
        unique_new = [n for n in new_data_raw if get_ord(n) not in existing_ids]
        
        full_data = existing_data + unique_new
        
        def get_chassis(r):
            k = next((key for key in r.keys() if str(key).strip().upper() in ["CHASSIS/SL NO.", "CHASSIS NO"]), None)
            return str(r.get(k, '')).strip() if k else ''
            
        full_data = [r for r in full_data if not get_chassis(r).startswith('B')]
        
        cr_dates = []
        for r in full_data:
            k_cr = next((key for key in r.keys() if str(key).strip().upper() in ["CR.DT", "CR DT"]), None)
            if k_cr and r.get(k_cr) and isinstance(r[k_cr], date):
                cr_dates.append(pd.to_datetime(r[k_cr]))
        latest_dt = max(cr_dates) if cr_dates else datetime.now()
        dt_str = latest_dt.strftime("%d/%m/%Y")
        dt_sheet = latest_dt.strftime("%d.%m.%y")
        dt_file = latest_dt.strftime("%d_%m_%Y")
        
        paycode_data = process_paycode_report(paycode_file)
        
        if ws.max_row > 1: ws.delete_rows(2, ws.max_row - 1)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        date_cols = ['Req. Delv. Dt', 'Cr.Dt', 'Closing Date', 'Date']
        numeric_cols = ['SNO.', 'TOTAL', 'PARTSTOTAL', 'ORDNO', 'CUSTNO']

        for r_idx, row_data in enumerate(full_data, start=2):
            
            ord_key = next((k for k in row_data.keys() if str(k).strip().upper() in ["ORD.NO.", "ORD NO"]), None)
            cr_key = next((k for k in row_data.keys() if str(k).strip().upper() in ["CR.DT", "CR DT"]), None)
            inv_key = next((k for k in row_data.keys() if str(k).strip().upper() in ["INVOICE NO.", "INVOICE NO", "INVOICE"]), None)
            dt_key = next((k for k in row_data.keys() if str(k).strip().upper() == "DATE"), None)
            cat_key = next((k for k in row_data.keys() if str(k).strip().upper() == "CATEGORY"), None)

            ord_no = str(row_data.get(ord_key, '')).split('.')[0].strip() if ord_key else ""
            creation_date = clean_date_string(row_data.get(cr_key)) if cr_key else None
            
            inv_no_raw = row_data.get(inv_key, '') if inv_key else ''
            existing_date_val = row_data.get(dt_key, '') if dt_key else ''
            inv_list = []
            final_date_val = clean_date_string(existing_date_val)
            has_cancelled_keyword = False

            if not pd.isna(inv_no_raw) and str(inv_no_raw).strip().lower() not in ['nan', 'none', '']:
                for i in str(inv_no_raw).split(','):
                    i_clean = i.strip()
                    if i_clean.upper() == 'CANCELLED':
                        has_cancelled_keyword = True
                    elif i_clean and not any(c in i_clean for c in ['-', ':']) and not inv_list:
                        inv_list.append(i_clean)
            
            has_correct_sequence = False
            has_invalid_sequence = False
            matched_inv = ""
            matched_dt = None

            if ord_no in paycode_data:
                for record_item in paycode_data[ord_no]:
                    inv = record_item["invoice"]
                    potential_date = record_item["date"]
                    
                    if inv and inv.upper() == 'CANCELLED':
                        has_cancelled_keyword = True
                        continue
                        
                    if isinstance(creation_date, date) and isinstance(potential_date, date):
                        if potential_date >= creation_date:
                            has_correct_sequence = True
                            if not matched_inv:
                                matched_inv = inv
                                matched_dt = potential_date
                        else:
                            has_invalid_sequence = True
                    else:
                        has_correct_sequence = True
                        if not matched_inv:
                            matched_inv = inv
                            matched_dt = potential_date

            if has_correct_sequence and matched_inv and not inv_list:
                inv_list.append(matched_inv)
                final_date_val = matched_dt

            final_inv_no = inv_list[0] if inv_list else ""
            has_inv = len(inv_list) > 0

            if has_inv:
                cat_val = "JOB CARD CLOSED"
                status = "Closed"
            else:
                orig_cat = str(row_data.get(cat_key, '') or '').strip() if cat_key else ""
                if orig_cat.upper() in ["CANCELLED", "CANCEL"]:
                    cat_val = "Cancelled"
                    status = "Closed"
                else:
                    cat_val = orig_cat
                    status = "Open"
                    if has_invalid_sequence and has_correct_sequence:
                        cat_val = "Cancelled"
                        status = "Closed"
                    elif has_invalid_sequence and not has_correct_sequence:
                        cat_val = ""
                        status = "Open"
                    elif has_cancelled_keyword:
                        cat_val = "Cancelled"
                        status = "Closed"
                    elif "CANCEL" in cat_val.upper() or "CLOSED" in cat_val.upper():
                        status = "Closed"

            for name, c_idx in col_map_ws.items():
                cell = ws.cell(row=r_idx, column=c_idx + 1)
                
                name_clean = str(name).strip().upper().replace(" ", "").replace(".", "")
                
                val = row_data.get(name)
                if val is None or str(val).strip() == "" or pd.isna(val) or str(val).strip().lower() == 'nan':
                    for k, v in row_data.items():
                        if str(k).strip().upper().replace(" ", "").replace(".", "") == name_clean:
                            val = v
                            break
                            
                if name_clean in ["SNO", "SLNO"]:
                    cell.value = r_idx - 1
                elif name_clean == "CATEGORY":
                    cell.value = cat_val
                elif name_clean == "STATUS":
                    cell.value = status
                elif name_clean in ["INVOICENO", "INVNO"]:
                    cell.number_format = '@'
                    cell.value = str(final_inv_no).strip() if final_inv_no else ""
                elif name_clean == "DATE":
                    cell.value = final_date_val
                elif name_clean == "REMARKS":
                    cell.value = apply_rich_remarks(str(val or '')) if pd.notna(val) and str(val).lower() != 'nan' else ""
                elif name_clean in ["PENDINGDAYS", "PENDING"]:
                    try: cell.value = int(float(str(val))) if val is not None and str(val).strip() != "" and str(val).lower() != 'nan' else ""
                    except: cell.value = val if pd.notna(val) and str(val).lower() != 'nan' else ""
                elif name_clean in ["DNAME", "DEALERCODE", "DEALERNAME"]:
                    # CRITICAL FIX: Extract D Name strictly using TEXTBEFORE("|", Dealer Name)
                    dlr_k = next((k for k in row_data.keys() if str(k).strip().upper().replace(" ", "").replace(".", "") == "DEALERNAME"), None)
                    dlr_v = str(row_data.get(dlr_k, '')).strip() if dlr_k else ""
                    
                    if "|" in dlr_v:
                        cell.value = to_numeric(dlr_v.split('|')[0].strip())
                    else:
                        cell.value = to_numeric(val) if pd.notna(val) and str(val).lower() != 'nan' else ""
                else:
                    if name_clean in numeric_cols:
                        cell.value = to_numeric(val)
                    else:
                        cell.value = val if pd.notna(val) and str(val).lower() != 'nan' else ""
                        
                if name_clean in ["REQDELVDT", "CRDT", "CLOSINGDATE", "DATE"] and cell.value:
                    cell.number_format = 'mm-dd-yy'

        if summary_file:
            for sname in list(wb.sheetnames):
                if "Summary WIP" in sname:
                    curr_ws = wb[sname]
                    title_prefix = "South Region WIP - Number of Job Cards as on " if "NO of JC" in sname else "South Region WIP as on "
                    curr_ws['B2'] = f"{title_prefix}{dt_str}"
                    base_name = "Summary WIP NO of JC" if "NO of JC" in sname else "Summary WIP"
                    curr_ws.title = f"{base_name} {dt_sheet}"
        
        if not summary_file:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(std_headers))}{len(full_data) + 1}"

        last_row = len(full_data) + 1
        cat_let = next((openpyxl.utils.get_column_letter(i+1) for k,i in col_map_ws.items() if str(k).strip().upper().replace(" ", "").replace(".", "") == "CATEGORY"), 'T')
        stat_let = next((openpyxl.utils.get_column_letter(i+1) for k,i in col_map_ws.items() if str(k).strip().upper().replace(" ", "").replace(".", "") == "STATUS"), 'W')
        inv_let = next((openpyxl.utils.get_column_letter(i+1) for k,i in col_map_ws.items() if str(k).strip().upper().replace(" ", "").replace(".", "") in ["INVOICENO", "INVNO"]), 'U')
        
        for r in range(2, last_row + 5):
            ws[f"{inv_let}{r}"].number_format = '@'

        ws.conditional_formatting.add(f'{cat_let}2:{cat_let}{last_row + 1000}', FormulaRule(formula=[f'OR({cat_let}2="JOB CARD CLOSED", {cat_let}2="Cancelled", {cat_let}2="CANCELLED")'], fill=green_fill, stopIfTrue=True))
        ws.conditional_formatting.add(f'{cat_let}2:{cat_let}{last_row + 1000}', FormulaRule(formula=[f'AND({cat_let}2<>"", {cat_let}2<>"JOB CARD CLOSED", {cat_let}2<>"Cancelled", {cat_let}2<>"CANCELLED")'], fill=blue_fill))
        
        red_text, green_text = Font(color="9C0006", bold=True), Font(color="006100", bold=True)
        ws.conditional_formatting.add(f'{stat_let}2:{stat_let}{last_row + 1000}', FormulaRule(formula=[f'{stat_let}2="Closed"'], fill=red_fill, font=red_text))
        ws.conditional_formatting.add(f'{stat_let}2:{stat_let}{last_row + 1000}', FormulaRule(formula=[f'{stat_let}2="Open"'], fill=green_fill, font=green_text))

        output = io.BytesIO()
        wb.save(output)
        
        if summary_file: st.success(f"Workbook updated based on Cr.Dt: {dt_str}")
        else: st.success(f"New Master Data generated from scratch for Date: {dt_str}")
        
        download_name = f"South_Region_WIP_Summary_{dt_file}.xlsx"
        st.download_button("📥 Download Final Report", output.getvalue(), file_name=download_name)
